#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from urllib.request import Request, urlopen

ROOT = Path(__file__).resolve().parents[1]
MARKET_DIR = ROOT / "data/market"
FRED_SERIES_URL = "https://fred.stlouisfed.org/data/MMMFFAA027N"
FRED_CSV_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id=MMMFFAA027N"
ICI_PAGE_URL = "https://www.ici.org/Research/FeesandExpenses"
SEC_URLS = [
    "https://www.sec.gov/divisions/investment/guidance/reservefundmmffaq.htm",
    "https://www.sec.gov/spotlight/reserve_primary_fund_investors/gardephe_opinion.pdf",
    "https://www.sec.gov/news/speech/2010/spch052610ajd.htm",
]
USER_AGENT = "KAFKA2306-investor2-us-mmf-collector/1.0 (+https://github.com/KAFKA2306/investor2)"


def fetch_bytes(url: str, timeout: int = 60) -> bytes:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    with urlopen(request, timeout=timeout) as response:
        return response.read()


def parse_fred_csv(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    rows = csv.DictReader(io.StringIO(text))
    if not rows.fieldnames or "observation_date" not in rows.fieldnames or "MMMFFAA027N" not in rows.fieldnames:
        raise AssertionError(f"unexpected FRED CSV columns: {rows.fieldnames}")
    records: list[dict[str, Any]] = []
    for row in rows:
        value = (row.get("MMMFFAA027N") or "").strip()
        if not value or value == ".":
            continue
        year = int((row.get("observation_date") or "")[:4])
        numeric = float(value)
        if numeric <= 0:
            continue
        records.append(
            {
                "type": "total_financial_assets",
                "year": year,
                "value_millions_usd": int(numeric) if numeric.is_integer() else numeric,
            }
        )
    if not records or records[0]["year"] > 1974 or records[-1]["year"] < 2025:
        raise AssertionError("FRED MMF history is unexpectedly incomplete")
    return records


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u2019", "'")).strip().lower()


def year_value(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and 1900 <= int(value) <= 2100:
        return int(value)
    text = normalize_text(value).replace("'", "")
    match = re.fullmatch(r"(?:19|20)?(\d{2})", text)
    if match:
        short = int(match.group(1))
        return 2000 + short if short < 70 else 1900 + short
    match = re.search(r"(19\d{2}|20\d{2})", text)
    return int(match.group(1)) if match else None


def numeric_value(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value).replace("%", "").replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_ici_workbook(raw: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse the ICI workbook") from exc

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    candidates: list[tuple[int, list[dict[str, Any]]]] = []

    for worksheet in workbook.worksheets:
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not rows:
            continue

        for index, row in enumerate(rows):
            normalized = [normalize_text(cell) for cell in row]
            gross_columns = [i for i, text in enumerate(normalized) if "gross yield" in text and "money market" in text]
            fed_columns = [i for i, text in enumerate(normalized) if "federal funds rate" in text]
            if not gross_columns or not fed_columns:
                continue
            year_columns = [i for i, text in enumerate(normalized) if text in {"year", "year-end", "year end"}]
            year_column = year_columns[0] if year_columns else 0
            gross_column = gross_columns[0]
            fed_column = fed_columns[0]
            parsed: list[dict[str, Any]] = []
            for data_row in rows[index + 1 :]:
                width = max(year_column, gross_column, fed_column)
                if len(data_row) <= width:
                    continue
                year = year_value(data_row[year_column])
                gross = numeric_value(data_row[gross_column])
                fed = numeric_value(data_row[fed_column])
                if year is None or gross is None or fed is None:
                    continue
                parsed.append(
                    {
                        "type": "taxable_mmf_yield",
                        "year": year,
                        "gross_yield_percent": round(gross, 6),
                        "federal_funds_rate_percent": round(fed, 6),
                    }
                )
            if len(parsed) >= 10:
                candidates.append((len(parsed), parsed))

        for year_row_index, row in enumerate(rows):
            positions = [(column, year_value(value)) for column, value in enumerate(row)]
            positions = [(column, year) for column, year in positions if year is not None]
            if len(positions) < 10:
                continue
            year_by_column = dict(positions)
            gross_row = None
            fed_row = None
            for nearby in rows[max(0, year_row_index - 12) : min(len(rows), year_row_index + 20)]:
                label = " ".join(normalize_text(value) for value in nearby if isinstance(value, str))
                if "gross yield" in label and "money market" in label:
                    gross_row = nearby
                if "federal funds rate" in label:
                    fed_row = nearby
            if gross_row is None or fed_row is None:
                continue
            parsed = []
            for column, year in sorted(year_by_column.items(), key=lambda item: item[1]):
                if column >= len(gross_row) or column >= len(fed_row):
                    continue
                gross = numeric_value(gross_row[column])
                fed = numeric_value(fed_row[column])
                if gross is None or fed is None:
                    continue
                parsed.append(
                    {
                        "type": "taxable_mmf_yield",
                        "year": year,
                        "gross_yield_percent": round(gross, 6),
                        "federal_funds_rate_percent": round(fed, 6),
                    }
                )
            if len(parsed) >= 10:
                candidates.append((len(parsed), parsed))

    if not candidates:
        raise AssertionError("could not locate ICI Figure 9 gross-yield and federal-funds-rate data")

    parsed = max(candidates, key=lambda item: (item[0], max(record["year"] for record in item[1])))[1]
    unique = {record["year"]: record for record in parsed if 2000 <= record["year"] <= 2100}
    result = [unique[year] for year in sorted(unique)]
    if result[0]["year"] > 2006 or result[-1]["year"] < 2025:
        raise AssertionError("ICI Figure 9 history is unexpectedly incomplete")
    return result


def latest_existing_artifact() -> dict[str, Any] | None:
    paths = sorted(MARKET_DIR.glob("us_mmf_history_*.json"))
    if not paths:
        return None
    return json.loads(paths[-1].read_text(encoding="utf-8"))


def discover_ici_workbook(existing: dict[str, Any] | None) -> tuple[str | None, bytes | None, list[str]]:
    urls: list[str] = []
    if existing:
        current = existing.get("sources", {}).get("taxable_mmf_yield", {}).get("data_url")
        if isinstance(current, str):
            urls.append(current)

    errors: list[str] = []
    try:
        page = fetch_bytes(ICI_PAGE_URL).decode("utf-8", errors="replace")
        hrefs = re.findall(r'href=["\']([^"\']+\.xlsx(?:\?[^"\']*)?)["\']', page, flags=re.IGNORECASE)
        for href in hrefs:
            url = urljoin(ICI_PAGE_URL, href)
            if url not in urls:
                urls.append(url)
    except Exception as exc:
        errors.append(f"{ICI_PAGE_URL}: discovery failed: {exc}")

    for url in urls[:12]:
        try:
            raw = fetch_bytes(url)
            parse_ici_workbook(raw)
            return url, raw, errors
        except Exception as exc:
            errors.append(f"{url}: {exc}")

    if existing:
        return None, None, errors
    raise AssertionError("no usable ICI supplemental workbook found and no verified baseline exists: " + " | ".join(errors))


def retained_ici_records(existing: dict[str, Any] | None, now: datetime) -> list[dict[str, Any]]:
    if not existing:
        raise AssertionError("ICI is unavailable and no verified ICI baseline exists")
    records = [record for record in existing.get("records", []) if record.get("type") == "taxable_mmf_yield"]
    if not records:
        raise AssertionError("verified baseline contains no ICI taxable-MMF yield records")
    latest_year = max(int(record["year"]) for record in records)
    required_year = now.year - 1 if now.month >= 4 else now.year - 2
    if latest_year < required_year:
        raise AssertionError(
            f"ICI annual series is stale: latest verified year={latest_year}, required>={required_year}; "
            "GitHub-hosted runner cannot retrieve the official ICI workbook, so a new primary-source snapshot is required"
        )
    return records


def infer_ici_pdf_url(data_url: str, existing: dict[str, Any] | None) -> str:
    match = re.search(r"/files/(\d{4})/(per\d+-\d+)-data\.xlsx", data_url)
    if match:
        year, stem = match.groups()
        existing_url = None if not existing else existing.get("sources", {}).get("taxable_mmf_yield", {}).get("pdf_url")
        if isinstance(existing_url, str) and stem in existing_url:
            return existing_url
        return f"https://www.ici.org/files/{year}/{stem}.pdf"
    return data_url


def break_the_buck_records() -> list[dict[str, Any]]:
    return [
        {
            "type": "break_the_buck",
            "year": 1994,
            "fund": "Community Bankers U.S. Government Money Market Fund",
            "nav_usd_per_share": None,
            "note": "SEC sources identify this as the only registered money market fund to break the buck before September 2008.",
        },
        {
            "type": "break_the_buck",
            "date": "2008-09-16",
            "year": 2008,
            "fund": "The Reserve Primary Fund",
            "nav_usd_per_share": 0.97,
            "lehman_exposure_usd": 785000000,
            "eventual_distribution_usd_per_share": 0.99,
        },
    ]


def comparable_payload(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "semantics": payload.get("semantics"),
        "records": payload.get("records"),
        "notes": payload.get("notes"),
    }


def compact_json(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, sort_keys=False, separators=(",", ":"))


def main() -> int:
    parser = argparse.ArgumentParser(description="Refresh the reusable U.S. money market fund history snapshot.")
    parser.add_argument("--check-only", action="store_true", help="Fetch and validate sources without writing a snapshot.")
    args = parser.parse_args()

    existing = latest_existing_artifact()
    now = datetime.now(timezone.utc)
    fred_records = parse_fred_csv(fetch_bytes(FRED_CSV_URL))
    ici_url, ici_raw, ici_errors = discover_ici_workbook(existing)
    if ici_raw is not None and ici_url is not None:
        ici_records = parse_ici_workbook(ici_raw)
        ici_access = {"status": "verified_live", "verified_at": now.isoformat(), "errors": ici_errors}
    else:
        ici_records = retained_ici_records(existing, now)
        previous_verified_at = None if not existing else existing.get("retrieved_at")
        ici_access = {
            "status": "retained_verified_snapshot",
            "verified_at": previous_verified_at,
            "errors": ici_errors,
        }
        ici_url = existing.get("sources", {}).get("taxable_mmf_yield", {}).get("data_url") if existing else None
    retrieved_at = now.isoformat()

    payload: dict[str, Any] = {
        "schema_version": "investor2.us-mmf-history.v1",
        "retrieved_at": retrieved_at,
        "component_status": {
            "total_financial_assets": {"status": "verified_live", "verified_at": retrieved_at},
            "taxable_mmf_yield": ici_access,
            "break_the_buck": {
                "status": "historical_fixed_evidence",
                "verified_at": existing.get("retrieved_at") if existing else retrieved_at,
            },
        },
        "semantics": {
            "total_financial_assets": "Annual year-end total financial assets; million USD.",
            "taxable_mmf_yield": "ICI Figure 9 year-end gross yield on taxable money market funds; percent.",
            "federal_funds_rate": "ICI Figure 9 year-end federal funds rate; percent.",
            "break_the_buck": "Historical events supported by the cited SEC sources; not a claim of exhaustive post-2008 event coverage.",
        },
        "sources": {
            "total_financial_assets": {
                "publisher": "Federal Reserve Bank of St. Louis / Board of Governors of the Federal Reserve System",
                "series_id": "MMMFFAA027N",
                "url": FRED_SERIES_URL,
                "note": "Official series is annual, end of period, millions of U.S. dollars. Zero values for 1945-1973 are omitted here; nonzero history starts in 1974.",
            },
            "taxable_mmf_yield": {
                "publisher": "Investment Company Institute",
                "figure": "Figure 9",
                "pdf_url": infer_ici_pdf_url(ici_url, existing) if isinstance(ici_url, str) else None,
                "data_url": ici_url,
            },
            "break_the_buck": {"publisher": "U.S. Securities and Exchange Commission", "urls": SEC_URLS},
        },
        "records": fred_records + ici_records + break_the_buck_records(),
        "notes": [
            "Money market funds are not principal-guaranteed deposits.",
            "Gross yield is not a guaranteed investor return.",
        ],
    }

    if existing and comparable_payload(existing) == comparable_payload(payload):
        print("US MMF history is unchanged; no snapshot created.")
        return 0

    if args.check_only:
        print(f"US MMF source validation passed; {len(payload['records'])} records would be materialized.")
        return 0

    artifact_bytes = compact_json(payload).encode("utf-8")
    artifact_sha256 = hashlib.sha256(artifact_bytes).hexdigest()
    date_part = retrieved_at[:10]
    artifact_path = MARKET_DIR / f"us_mmf_history_{date_part}_{artifact_sha256[:12]}.json"
    artifact_path.parent.mkdir(parents=True, exist_ok=True)
    artifact_path.write_bytes(artifact_bytes)

    sys.path.insert(0, str(ROOT / "scripts"))
    from snapshot_store import append_entry, build_entry, load_registry

    source_urls = [FRED_SERIES_URL, FRED_CSV_URL, ICI_PAGE_URL, *SEC_URLS]
    if isinstance(ici_url, str):
        source_urls.extend([ici_url, infer_ici_pdf_url(ici_url, existing)])
    source_urls = list(dict.fromkeys(source_urls))
    entry = build_entry(
        root=ROOT,
        registry=load_registry(),
        dataset_id="us_mmf_history",
        reuse_key="us/mmf/history/assets-yield-break-the-buck",
        artifact_path=artifact_path.relative_to(ROOT).as_posix(),
        source="public_web_research",
        source_kind="official_web",
        observed_at=retrieved_at,
        schema_version="investor2.us-mmf-history.v1",
        provenance={
            "tool": "GitHub Actions",
            "operation": "Refresh verified U.S. money market fund history from official public sources.",
            "query_or_scope": "Annual U.S. MMF total financial assets; ICI Figure 9 taxable MMF gross yields and federal funds rates; SEC-supported 1994 and 2008 break-the-buck events.",
            "retrieved_at": retrieved_at,
            "source_urls": source_urls,
        },
    )
    append_entry(entry)
    print(json.dumps({"artifact_path": entry["artifact_path"], "snapshot_id": entry["snapshot_id"], "artifact_sha256": artifact_sha256}, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
